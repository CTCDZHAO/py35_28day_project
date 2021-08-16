import openpyxl
import os

"""
数据的写入：
      1、文件名：
      2、表单：
      3、写入行
      4、写入列
      
"""
class HandlExcel:
    def __init__(self,filename,sheetname):
        """

        :param filename: excel(文件名)
        :param sheetname: 表单名
        """
        self.filename=filename
        self.sheetname=sheetname

    def read_data(self):
        """
        数据读取

        """
        workbook = openpyxl.load_workbook(self.filename)
        she = workbook[self.sheetname]
        res = list(she.rows)
        title = [i.value for i in res[0]]
        case = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            case.append(dic)
        return case
    def write_data(self,row,column,value):
        '''数据写入方法
        :param row:写入行
        :param column:写入列
        :param value:写入值
        :return:
        '''
        workbook = openpyxl.load_workbook(self.filename)
        she = workbook[self.sheetname]
        # excel中写入数据
        she.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


# def readexcel():
#     workbook=openpyxl.load_workbook(r'D:\PycharmProjects\ningmeng35\py35_14day\task_13day\test_data\cases.xlsx')
#     she=workbook['Sheet1']
#     res=list(she.rows)
#     title=[i.value for i in res[0]]
#     case=[]
#     for item in res[1:]:
#         data=[i.value for i in item]
#         dic=dict(zip(title,data))
#         case.append(dic)
#     return case
if __name__ == '__main__':
    excel=HandlExcel(r'D:\PycharmProjects\ningmeng35\py35_15day\test001.xlsx','register')
    case=excel.read_data()
    print(case)

